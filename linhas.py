import openpyxl
import math

def split_excel(filename, num_rows):
    # Carrega o arquivo Excel
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    
    # Obtém o número total de linhas
    total_rows = sheet.max_row
    
    # Calcula o número de arquivos necessários
    num_files = math.ceil(total_rows / num_rows)
    
    # Itera sobre cada arquivo
    for file_num in range(num_files):
        # Cria um novo workbook para cada arquivo
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        
        # Copia as linhas para o novo workbook
        start_row = file_num * num_rows + 1
        end_row = min((file_num + 1) * num_rows, total_rows)
        
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
            new_sheet.append(row)
        
        # Salva o novo workbook com um nome de arquivo único
        new_workbook.save(f'output_{file_num + 1}.xlsx')

# Uso do script
input_file = '/Users/williamduarte/Library/Mobile Documents/com~apple~CloudDocs/Documents/Indicadores/contatos talos sapataria nova (1).xlsx'
split_excel(input_file, num_rows=10000)